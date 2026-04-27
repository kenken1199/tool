import platform
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.font_manager as fm
import numpy as np
from scipy import stats
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime
from io import BytesIO
import os
import sys


# =========================
# ■ 定数
# =========================
MIN_OK_COUNT = 2  # 統計分析に必要な最小OKデータ数


def _setup_japanese_font():
    candidates = {
        "Windows": ["MS Gothic", "Yu Gothic", "Meiryo"],
        "Darwin":  ["Hiragino Sans", "Hiragino Maru Gothic Pro", "AppleGothic"],
        "Linux":   ["Noto Sans CJK JP", "IPAGothic", "IPAPGothic"],
    }.get(platform.system(), [])
    available = {f.name for f in fm.fontManager.ttflist}
    for font in candidates:
        if font in available:
            plt.rcParams["font.family"] = font
            return

_setup_japanese_font()

app_root = None
csv_file_to_process = None


# =========================
# ■ ロットプレビューダイアログ
# =========================
class LotPreviewDialog(tk.Toplevel):

    def __init__(self, parent, df, hinshoku_num=None):
        super().__init__(parent)
        self.title("ロット分割プレビュー")
        self.result = None
        self.df = df.copy()
        self.hinshoku_num = hinshoku_num
        self.resizable(True, True)
        self.grab_set()

        # --- しきい値入力 ---
        frame_top = ttk.Frame(self, padding=10)
        frame_top.pack(fill="x")

        ttk.Label(frame_top, text="分割しきい値（分）:").pack(side="left")
        self.threshold_var = tk.IntVar(value=30)
        ttk.Spinbox(
            frame_top, from_=1, to=480,
            textvariable=self.threshold_var, width=6
        ).pack(side="left", padx=5)
        ttk.Button(frame_top, text="更新", command=self._update_preview).pack(side="left", padx=5)

        self.lot_label_var = tk.StringVar()
        ttk.Label(frame_top, textvariable=self.lot_label_var, foreground="navy").pack(side="left", padx=15)

        # --- Treeview ---
        frame_tree = ttk.Frame(self, padding=10)
        frame_tree.pack(fill="both", expand=True)

        cols = ("品種番号", "ロット", "開始時刻", "終了時刻", "総件数", "OKデータ件数", "状態")
        self.tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=10)
        col_widths = {"品種番号": 80, "ロット": 70, "開始時刻": 150, "終了時刻": 150,
                      "総件数": 80, "OKデータ件数": 110, "状態": 110}
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=col_widths[col])

        # 警告色タグ（OKデータ不足のロットを赤系で強調）
        self.tree.tag_configure("skip", background="#FFE4E1", foreground="#9C0006")

        scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # --- 注意書きラベル ---
        self.warn_label_var = tk.StringVar()
        warn_label = ttk.Label(self, textvariable=self.warn_label_var,
                               foreground="#9C0006", padding=(10, 0))
        warn_label.pack(fill="x")

        # --- ボタン ---
        frame_btn = ttk.Frame(self, padding=(10, 5, 10, 10))
        frame_btn.pack(fill="x")
        ttk.Button(frame_btn, text="この分割でOK", command=self._on_ok).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="ロット分割しない", command=self._on_no_split).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="手動分割する", command=self._on_manual).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="キャンセル", command=self._on_cancel).pack(side="right", padx=5)

        self._update_preview()

        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _compute_lots(self, threshold_min):
        df = self.df.sort_values("日付時刻").copy()
        df["時間差(分)"] = df["日付時刻"].diff().dt.total_seconds() / 60
        df["ロット"] = (df["時間差(分)"] > threshold_min).cumsum() + 1
        return df

    def _update_preview(self):
        threshold = self.threshold_var.get()
        df = self._compute_lots(threshold)

        for item in self.tree.get_children():
            self.tree.delete(item)

        lot_count = df["ロット"].nunique()
        skip_count = 0

        hinshoku_display = str(self.hinshoku_num) if self.hinshoku_num is not None else "-"

        for lot, group in df.groupby("ロット"):
            start = group["日付時刻"].min()
            end = group["日付時刻"].max()
            total = len(group)
            ok_count = int((group["ランクコード"] == "2").sum())

            if ok_count < MIN_OK_COUNT:
                state = "⚠ スキップ予定"
                tags = ("skip",)
                skip_count += 1
            else:
                state = "✓ 分析対象"
                tags = ()

            self.tree.insert("", "end", tags=tags, values=(
                hinshoku_display,
                f"ロット{lot}",
                start.strftime("%Y-%m-%d %H:%M") if pd.notna(start) else "-",
                end.strftime("%Y-%m-%d %H:%M") if pd.notna(end) else "-",
                total,
                ok_count,
                state,
            ))

        # 上部ラベル
        if skip_count > 0:
            self.lot_label_var.set(
                f"→ {lot_count} ロット検出（うち {skip_count} ロットはOKデータ不足でスキップ予定）"
            )
            self.warn_label_var.set(
                f"※ OKデータが {MIN_OK_COUNT} 件未満のロットは統計計算ができないため、"
                f"Excelファイルは作成されません。しきい値を変更してロットを統合することも検討してください。"
            )
        else:
            self.lot_label_var.set(f"→ {lot_count} ロット検出")
            self.warn_label_var.set("")

        self._df_with_lots = df

    def _on_ok(self):
        self.result = ("ok", self._df_with_lots)
        self.destroy()

    def _on_no_split(self):
        df = self.df.copy()
        df["ロット"] = 1
        self.result = ("ok", df)
        self.destroy()

    def _on_manual(self):
        self.result = ("manual", None)
        self.destroy()

    def _on_cancel(self):
        self.result = ("cancel", None)
        self.destroy()


# =========================
# ■ CSV正規化
# =========================
_ISHIDA_RANK_VALUES = {"正量", "軽量", "過量"}


def _read_csv(file, **kwargs):
    for enc in ("cp932", "utf-8-sig"):
        try:
            return pd.read_csv(file, encoding=enc, **kwargs)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"CSVのエンコーディングを判別できませんでした: {os.path.basename(file)}")


def normalize_columns(file):

    df = _read_csv(file)
    df.columns = df.columns.str.replace("　", "").str.replace(" ", "").str.strip()
    df = df.loc[:, ~df.columns.duplicated()]
    cols = df.columns.tolist()

    # ===== アンリツ =====
    if any("測定値" in col for col in cols) and any("ランクコード" in col for col in cols):

        rename_map = {}
        for col in cols:
            if col == "測定値出力No.":
                rename_map[col] = "測定値出力No."
            elif col == "測定値(g)":
                rename_map[col] = "測定値(g)"
            elif "ランクコード" in col:
                rename_map[col] = "ランクコード"
            elif "日付時刻" in col:
                rename_map[col] = "日付時刻"

        df = df.rename(columns=rename_map)

        if "日付時刻" not in df.columns:
            if "日付" in df.columns and "時刻" in df.columns:
                df["日付時刻"] = pd.to_datetime(
                    df["日付"].astype(str) + " " + df["時刻"].astype(str),
                    errors="coerce"
                )

        hinshoku_num = None
        if "品種" in df.columns:
            vals = pd.to_numeric(df["品種"], errors="coerce").dropna()
            if len(vals) > 0:
                hinshoku_num = int(vals.iloc[0])

        df["メーカー"] = "アンリツ"
        df = df[["測定値出力No.", "日付時刻", "測定値(g)", "ランクコード", "メーカー"]].copy()
        df["測定値出力No."] = pd.to_numeric(df["測定値出力No."], errors="coerce")
        df["測定値(g)"] = pd.to_numeric(df["測定値(g)"], errors="coerce")
        df["ランクコード"] = df["ランクコード"].astype(str).str.strip()
        df["日付時刻"] = pd.to_datetime(df["日付時刻"], errors="coerce")
        return df, hinshoku_num

    # ===== イシダ判定 =====
    df_ishida = _read_csv(file, skiprows=10)
    df_ishida.columns = df_ishida.columns.str.replace("　", "").str.replace(" ", "").str.strip()

    is_ishida = (
        len(df_ishida.columns) >= 6
        and df_ishida.iloc[:, 5].dropna().astype(str).isin(_ISHIDA_RANK_VALUES).any()
    )
    if not is_ishida:
        raise ValueError(
            f"未対応のCSVフォーマットです。アンリツまたはイシダ形式のCSVを選択してください。\n"
            f"ファイル: {os.path.basename(file)}"
        )

    hinshoku_num = None
    if "予約番号" in df_ishida.columns:
        vals = pd.to_numeric(df_ishida["予約番号"], errors="coerce").dropna()
        if len(vals) > 0:
            hinshoku_num = int(vals.iloc[0])
    elif len(df_ishida.columns) >= 4:
        vals = pd.to_numeric(df_ishida.iloc[:, 3], errors="coerce").dropna()
        if len(vals) > 0:
            hinshoku_num = int(vals.iloc[0])

    df_ishida = df_ishida.iloc[:, [0, 1, 4, 5]].copy()
    df_ishida.columns = ["日付", "時刻", "測定値(g)", "判定"]

    df_ishida["日付時刻"] = pd.to_datetime(
        df_ishida["日付"].astype(str) + " " + df_ishida["時刻"].astype(str),
        errors="coerce"
    )
    df_ishida["測定値出力No."] = range(1, len(df_ishida) + 1)
    rank_map = {"正量": "2", "軽量": "1", "過量": "E"}
    df_ishida["ランクコード"] = df_ishida["判定"].map(rank_map)
    df_ishida["メーカー"] = "イシダ"

    return df_ishida[["測定値出力No.", "日付時刻", "測定値(g)", "ランクコード", "メーカー"]], hinshoku_num


# =========================
# ■ 分析
# =========================
def analyze(data):

    data = np.asarray(data).astype(float).ravel()
    data = data[~np.isnan(data)]

    n = len(data)
    if n < 2:
        return None, None, (None, None), None, None, None, None

    mean = float(np.mean(data))
    std = float(np.std(data, ddof=1))

    t_value = stats.t.ppf(0.975, df=n - 1)
    margin = t_value * std / np.sqrt(n)

    max1 = float(np.max(data))
    min1 = float(np.min(data))
    lower = mean - 3 * std
    upper = mean + 3 * std

    return mean, std, (mean - margin, mean + margin), max1, min1, lower, upper


# =========================
# ■ Excel出力
# =========================
def _create_report_sheet(wb, df_ok, mean, std, ci, max1, min1, lower, upper,
                          outliers_df, img_hist_bytes, img_series_bytes, rank_counts,
                          total_count, original_ok_count, hinshoku_num, date_str, lot):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    ws = wb.create_sheet("分析レポート", 0)

    # スタイル
    title_font  = Font(bold=True, size=16, color="FFFFFFFF")
    title_fill  = PatternFill(start_color="FF1F3864", fill_type="solid")
    sec_font    = Font(bold=True, size=10, color="FFFFFFFF")
    sec_fill    = PatternFill(start_color="FF4472C4", fill_type="solid")
    label_font  = Font(bold=True, size=10)
    label_fill  = PatternFill(start_color="FFD9E1F2", fill_type="solid")
    val_fill_e  = PatternFill(start_color="FFFFFFFF", fill_type="solid")
    val_fill_o  = PatternFill(start_color="FFEFF3FB", fill_type="solid")
    warn_fill   = PatternFill(start_color="FFFFF2CC", fill_type="solid")
    warn_font   = Font(bold=True, size=10, color="FF7F6000")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", indent=1)

    # 列幅
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 13
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 10.4  # D〜J合計 ≈ 13.5cm

    ng_count    = total_count - original_ok_count
    defect_rate = (ng_count / total_count * 100) if total_count > 0 else 0.0

    # タイトル行
    if hinshoku_num is not None and date_str:
        title = f"{date_str}製造   品種番号 {hinshoku_num}   ロット{lot}   分析レポート"
    else:
        title = f"ロット{lot}   分析レポート"

    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font      = title_font
    ws["A1"].fill      = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells("K1:L1")
    ws["K1"] = datetime.datetime.now().strftime("%Y年%m月%d日")
    ws["K1"].font      = Font(bold=True, size=11, color="FFFFFFFF")
    ws["K1"].fill      = title_fill
    ws["K1"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # 統計セクションヘッダー
    ws.merge_cells("A3:B3")
    ws["A3"] = "■ 基本統計"
    ws["A3"].font      = sec_font
    ws["A3"].fill      = sec_fill
    ws["A3"].alignment = center
    ws.row_dimensions[3].height = 16

    # 統計テーブル定義: (ラベル, 値, 書式, 警告フラグ)
    stats_rows = [
        ("全数",             total_count,                              "0",     False),
        ("OK数",             original_ok_count,                       "0",     False),
        ("NG数",             ng_count,                                 "0",     ng_count > 0),
        ("不良率 (%)",        defect_rate,                              "0.00",  ng_count > 0),
        (None, None, None, False),
        ("平均 (g)",          mean,                                    "0.0", False),
        ("標準偏差 (g)",       std,                                    "0.000", False),
        ("OKデータ件数",      len(df_ok),                              "0",     False),
        ("Max (g)",          max1,                                    "0.0", False),
        ("Min (g)",          min1,                                    "0.0", False),
        ("下限 −3σ (g)",     lower,                                   "0.0", False),
        ("上限 +3σ (g)",     upper,                                   "0.0", False),
        (None, None, None, False),
        ("95% CI 下限 (g)",  ci[0] if ci[0] is not None else "−",    "0.0", False),
        ("95% CI 上限 (g)",  ci[1] if ci[1] is not None else "−",    "0.0", False),
        ("外れ値件数",         len(outliers_df),                        "0",     len(outliers_df) > 0),
    ]

    data_row = 4
    stripe   = 0
    for label, value, fmt, warn in stats_rows:
        ws.row_dimensions[data_row].height = 5 if label is None else 14
        if label is None:
            data_row += 1
            continue
        vfill = warn_fill if warn else (val_fill_e if stripe % 2 == 0 else val_fill_o)
        vfont = warn_font if warn else Font(size=10)

        ws[f"A{data_row}"] = label
        ws[f"A{data_row}"].font      = label_font
        ws[f"A{data_row}"].fill      = label_fill
        ws[f"A{data_row}"].border    = thin
        ws[f"A{data_row}"].alignment = left

        ws[f"B{data_row}"] = value
        ws[f"B{data_row}"].font          = vfont
        ws[f"B{data_row}"].fill          = vfill
        ws[f"B{data_row}"].border        = thin
        ws[f"B{data_row}"].alignment     = center
        ws[f"B{data_row}"].number_format = fmt

        data_row += 1
        stripe   += 1

    ws.row_dimensions[data_row].height = 7  # 統計〜ランク間ギャップ（行3〜25計≈10.5cm調整）

    # ランクコード集計ミニテーブル
    rank_row = data_row + 1
    ws.merge_cells(f"A{rank_row}:B{rank_row}")
    ws[f"A{rank_row}"] = "■ ランクコード集計"
    ws[f"A{rank_row}"].font      = sec_font
    ws[f"A{rank_row}"].fill      = sec_fill
    ws[f"A{rank_row}"].alignment = center
    ws.row_dimensions[rank_row].height = 16

    rank_header_row = rank_row + 1
    for ci_idx, col_name in enumerate([ "内容", "件数"]):
        col_letter = ["A", "B", "C"][ci_idx] if ci_idx < 3 else "A"
        cell = ws.cell(row=rank_header_row, column=ci_idx + 1, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFFFF")
        cell.fill      = PatternFill(start_color="FF4472C4", fill_type="solid")
        cell.border    = thin
        cell.alignment = center
    ws.row_dimensions[rank_header_row].height = 14

    for r_idx, row_data in rank_counts[["内容", "件数"]].iterrows():
        r = rank_header_row + 1 + r_idx
        rfill = val_fill_e if r_idx % 2 == 0 else val_fill_o
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.fill      = rfill
            cell.border    = thin
            cell.alignment = center
        ws.row_dimensions[r].height = 13

    # ヒストグラム (D3:J25 に TwoCellAnchor で固定 ≈ 13.5cm × 10.5cm)
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    img1 = Image(BytesIO(img_hist_bytes))
    img1.width  = 510  # fallback: 13.5cm
    img1.height = 397  # fallback: 10.5cm
    anchor1 = TwoCellAnchor(editAs="twoCell")
    anchor1._from = AnchorMarker(col=3, colOff=0, row=2,  rowOff=0)  # D3  (0-indexed)
    anchor1.to    = AnchorMarker(col=9, colOff=0, row=26, rowOff=0)  # J25 (0-indexed)
    ws.add_image(img1, anchor1)

    # 時系列チャート (下段)
    series_row = max(data_row, rank_header_row + len(rank_counts) + 1) + 2
    ws.merge_cells(f"A{series_row}:L{series_row}")
    ws[f"A{series_row}"] = "■ 時系列チャート"
    ws[f"A{series_row}"].font      = sec_font
    ws[f"A{series_row}"].fill      = sec_fill
    ws[f"A{series_row}"].alignment = center
    ws.row_dimensions[series_row].height = 16

    img2 = Image(BytesIO(img_series_bytes))
    img2.width  = 907  # 24cm（紙幅いっぱい→fit-to-pageで左右対称に印刷）
    img2.height = 378  # 10cm
    ws.add_image(img2, f"A{series_row + 1}")

    # A4縦・1ページ印刷設定
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def save_to_excel(df_ok, mean, std, ci, max1, min1, lower, upper,
                  outliers_df, img_hist, img_series, rank_counts, filename, lot,
                  total_count=0, original_ok_count=0, hinshoku_num=None, date_str=None):

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    red_fill = PatternFill(start_color="FFFF0000", fill_type="solid")
    header_fill = PatternFill(start_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    result_fill = PatternFill(start_color="FFE7E6E6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ng_count = total_count - original_ok_count
    defect_rate = (ng_count / total_count * 100) if total_count > 0 else 0.0

    result_df = pd.DataFrame({
        "項目": ["全数", "OK数", "NG数", "不良率(%)", "平均", "標準偏差", "データ数", "Max", "Min", "下限(-3σ)", "上限(+3σ)", "95%CI 下限", "95%CI 上限"],
        "値": [total_count, original_ok_count, ng_count, defect_rate, mean, std, len(df_ok), max1, min1, lower, upper,
               ci[0] if ci[0] is not None else None, ci[1] if ci[1] is not None else None]
    })

    if hinshoku_num is not None and date_str:
        title_str = f"{date_str}製造 品種番号{hinshoku_num} 統計結果"
    else:
        title_str = None

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        result_df.to_excel(writer, sheet_name="統計結果", index=False, startrow=1 if title_str else 0)
        df_ok[["測定値出力No.", "日付時刻", "測定値(g)"]].to_excel(writer, sheet_name="OKデータ", index=False)
        outliers_df[["測定値出力No.", "日付時刻", "測定値(g)"]].to_excel(writer, sheet_name="外れ値", index=False)
        rank_counts.to_excel(writer, sheet_name="ランクコード集計", index=False)

        wb = writer.book

        # ===== 統計結果シート =====
        ws_result = wb["統計結果"]
        ws_result.column_dimensions["A"].width = 20
        ws_result.column_dimensions["B"].width = 20

        if title_str:
            ws_result["A1"] = title_str
            ws_result["A1"].font = Font(bold=True, size=12)
            ws_result["A1"].alignment = center_align
            ws_result.merge_cells("A1:B1")
            header_row = 2
            data_start_row = 3
        else:
            header_row = 1
            data_start_row = 2

        for cell in ws_result[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        for row in ws_result.iter_rows(min_row=data_start_row, max_row=ws_result.max_row):
            for cell in row:
                cell.fill = result_fill
                cell.border = border
                cell.alignment = center_align
                if cell.column == 2:
                    item = ws_result.cell(row=cell.row, column=1).value
                    if item == "標準偏差":
                        cell.number_format = "0.000"
                    elif item in ("データ数", "全数", "OK数", "NG数"):
                        cell.number_format = "0"
                    elif item == "不良率(%)":
                        cell.number_format = "0.00"
                    else:
                        cell.number_format = "0.0"

        # ===== ランクコード集計シート =====
        ws_rank = wb["ランクコード集計"]
        ws_rank.column_dimensions["A"].width = 15
        ws_rank.column_dimensions["B"].width = 15
        ws_rank.column_dimensions["C"].width = 20
        for cell in ws_rank[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        for row in ws_rank.iter_rows(min_row=2, max_row=ws_rank.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = center_align
                if row[0].row % 2 == 0:
                    cell.fill = PatternFill(start_color="FFF2F2F2", fill_type="solid")

        # ===== OKデータシート =====
        ws_ok = wb["OKデータ"]
        ws_ok.column_dimensions["A"].width = 18
        ws_ok.column_dimensions["B"].width = 22
        ws_ok.column_dimensions["C"].width = 15
        for cell in ws_ok[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        for row in ws_ok.iter_rows(min_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
        for row in ws_ok.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        ws_ok.auto_filter.ref = f"A1:C{ws_ok.max_row}"
        outlier_ids = set(outliers_df["測定値出力No."].values)
        for row in ws_ok.iter_rows(min_row=2, max_row=ws_ok.max_row):
            if row[0].value in outlier_ids:
                for cell in row:
                    cell.fill = red_fill

        # ===== 外れ値シート =====
        ws_out = wb["外れ値"]
        ws_out.column_dimensions["A"].width = 18
        ws_out.column_dimensions["B"].width = 22
        ws_out.column_dimensions["C"].width = 15
        for cell in ws_out[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        for row in ws_out.iter_rows(min_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
        for row in ws_out.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        ws_out.auto_filter.ref = f"A1:C{ws_out.max_row}"

        # ===== グラフシート =====
        img_hist_bytes   = img_hist.getvalue()
        img_series_bytes = img_series.getvalue()

        ws_hist = wb.create_sheet("ヒストグラム")
        ws_hist.add_image(Image(BytesIO(img_hist_bytes)), "A1")

        ws_series = wb.create_sheet("時系列チャート")
        ws_series.add_image(Image(BytesIO(img_series_bytes)), "A1")

        # ===== 分析レポートシート (先頭に挿入) =====
        _create_report_sheet(
            wb, df_ok, mean, std, ci, max1, min1, lower, upper,
            outliers_df, img_hist_bytes, img_series_bytes, rank_counts,
            total_count, original_ok_count, hinshoku_num, date_str, lot
        )


# =========================
# ■ ロット処理
# =========================
def process_lot(group, lot, save_dir, hinshoku_num=None):
    """
    1ロット分の分析を行いExcelを出力する。
    戻り値:
        ("ok",   ok_count) … 正常に作成
        ("skip", ok_count) … OKデータ不足によりスキップ
    """

    rank_map = {"2": "OK", "1": "軽量", "E": "過量", "0": "２個乗り"}

    rank_counts = group["ランクコード"].value_counts().reset_index()
    rank_counts.columns = ["ランクコード", "件数"]
    rank_counts["内容"] = rank_counts["ランクコード"].map(rank_map)
    rank_counts = rank_counts[["ランクコード", "内容", "件数"]]

    total_count = len(group)
    original_ok_count = int((group["ランクコード"] == "2").sum())

    df_ok = group[group["ランクコード"] == "2"].copy()

    data = pd.to_numeric(df_ok["測定値(g)"], errors="coerce")
    df_ok = df_ok.loc[data.notna()].copy()
    data = data.loc[data.notna()]
    data = np.asarray(data).ravel()

    if len(data) < MIN_OK_COUNT:
        return ("skip", len(data))

    mean, std, ci, max1, min1, lower, upper = analyze(data)

    outliers_df = df_ok[(df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)]

    lot_date = group["日付時刻"].dropna().min()
    if pd.notna(lot_date) and hinshoku_num is not None:
        date_str = f"{lot_date.year}/{lot_date.month}/{lot_date.day}"
        chart_prefix = f"{date_str}製造 品種番号{hinshoku_num} "
        date_str_safe = date_str.replace("/", "-")
    else:
        date_str = None
        chart_prefix = ""
        date_str_safe = None

    # ===== ヒストグラム =====
    fig1, ax1 = plt.subplots(figsize=(10, 6))
    ax1.hist(data, bins=30, edgecolor="black", alpha=0.7)
    ax1.axvline(mean, color="red", linestyle="-", linewidth=2, label=f"平均: {mean:.2f}")
    ax1.axvline(lower, color="orange", linestyle="--", linewidth=2, label=f"下限(-3σ): {lower:.2f}")
    ax1.axvline(upper, color="orange", linestyle="--", linewidth=2, label=f"上限(+3σ): {upper:.2f}")
    ax1.set_title(f"{chart_prefix}測定値の分布（n={len(data)}）", fontsize=14, fontweight="bold")
    ax1.set_xlabel("測定値(g)", fontsize=12)
    ax1.set_ylabel("頻度", fontsize=12)
    ax1.legend(fontsize=10)
    ax1.grid(True, alpha=0.3)
    fig1.tight_layout()

    img_hist = BytesIO()
    fig1.savefig(img_hist, format="png", dpi=100, bbox_inches="tight")
    img_hist.seek(0)
    plt.close(fig1)

    # ===== 時系列チャート =====
    fig2, ax2 = plt.subplots(figsize=(12, 5))

    y_vals = df_ok["測定値(g)"].values
    outlier_mask = (df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)
    has_datetime = df_ok["日付時刻"].notna().any()

    if has_datetime:
        x_all = df_ok["日付時刻"]
        x_ok  = df_ok.loc[~outlier_mask, "日付時刻"]
        x_out = df_ok.loc[outlier_mask,  "日付時刻"]
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax2.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax2.set_xlabel("時刻", fontsize=12)
    else:
        x_all = np.arange(1, len(df_ok) + 1)
        x_ok  = x_all[~outlier_mask.values]
        x_out = x_all[outlier_mask.values]
        ax2.set_xlabel("測定順序", fontsize=12)

    y_ok  = y_vals[~outlier_mask.values]
    y_out = y_vals[outlier_mask.values]

    ax2.plot(x_all, y_vals, color="steelblue", linewidth=0.6, alpha=0.4, zorder=1)
    ax2.scatter(x_ok, y_ok, color="steelblue", s=18, alpha=0.8, zorder=2, label="OK")
    if len(x_out) > 0:
        ax2.scatter(x_out, y_out, color="red", s=50, marker="x",
                    linewidths=2, zorder=3, label=f"外れ値 ({len(x_out)}件)")

    ax2.axhline(mean,  color="red",    linewidth=1.5, linestyle="-",  label=f"平均: {mean:.2f}")
    ax2.axhline(upper, color="orange", linewidth=1.5, linestyle="--", label=f"+3σ: {upper:.2f}")
    ax2.axhline(lower, color="orange", linewidth=1.5, linestyle="--", label=f"-3σ: {lower:.2f}")

    ax2.set_title(f"{chart_prefix}時系列チャート（n={len(df_ok)}）", fontsize=14, fontweight="bold")
    ax2.set_ylabel("測定値(g)", fontsize=12)
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)
    fig2.tight_layout()

    img_series = BytesIO()
    fig2.savefig(img_series, format="png", dpi=100, bbox_inches="tight")
    img_series.seek(0)
    plt.close(fig2)

    if date_str_safe and hinshoku_num is not None:
        filename = os.path.join(
            save_dir,
            f"分析結果_{date_str_safe}製造_品種番号{hinshoku_num} ロット{lot}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
        )
    else:
        filename = os.path.join(
            save_dir,
            f"分析結果_ロット{lot}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
        )

    save_to_excel(df_ok, mean, std, ci, max1, min1,
                  lower, upper, outliers_df,
                  img_hist, img_series, rank_counts, filename, lot,
                  total_count=total_count, original_ok_count=original_ok_count,
                  hinshoku_num=hinshoku_num, date_str=date_str)

    return ("ok", len(data))


# =========================
# ■ ファイル処理
# =========================
def process_file(file):
    try:
        if not file or not os.path.exists(file):
            messagebox.showerror("エラー", "ファイルが見つかりません")
            return

        save_dir = os.path.dirname(file)
        df, hinshoku_num = normalize_columns(file)

        dialog = LotPreviewDialog(app_root, df, hinshoku_num)
        app_root.wait_window(dialog)

        if dialog.result is None or dialog.result[0] == "cancel":
            return
        if dialog.result[0] == "manual":
            messagebox.showwarning("中止", "CSVを手動分割してください")
            return

        df = dialog.result[1]

        # 結果集計
        created_lots = []   # [(lot, ok_count), ...]
        skipped_lots = []   # [(lot, ok_count, total_count), ...]

        for lot, group in df.groupby("ロット"):
            total = len(group)
            status, ok_count = process_lot(group, lot, save_dir, hinshoku_num)
            if status == "ok":
                created_lots.append((lot, ok_count))
            else:
                skipped_lots.append((lot, ok_count, total))

        # ===== 完了メッセージ =====
        if not created_lots and skipped_lots:
            # すべてスキップされた異常ケース
            msg = "Excelファイルは作成されませんでした。\n\n"
            msg += "すべてのロットでOKデータが不足しています:\n"
            for lot, ok, total in skipped_lots:
                msg += f"  ・ロット{lot}: 総{total}件 / OK{ok}件\n"
            msg += "\nしきい値を変更するか、CSVの内容をご確認ください。"
            messagebox.showerror("作成失敗", msg)

        elif skipped_lots:
            # 一部スキップ
            msg = f"Excel作成完了\n作成: {len(created_lots)}ファイル\n\n"
            msg += "⚠ 以下のロットはOKデータ不足のためスキップしました:\n"
            for lot, ok, total in skipped_lots:
                msg += f"  ・ロット{lot}: 総{total}件 / OK{ok}件\n"
            msg += f"\n（OKデータが {MIN_OK_COUNT} 件未満のロットは統計計算ができません）"
            messagebox.showwarning("完了（一部スキップ）", msg)

        else:
            # 全件成功
            messagebox.showinfo(
                "完了",
                f"Excel作成完了\n{len(created_lots)}ファイルを作成しました。"
            )

    except Exception as e:
        messagebox.showerror("エラー", str(e))


# =========================
# ■ メイン
# =========================
def run():
    file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if file:
        process_file(file)


def on_closing():
    if app_root:
        app_root.destroy()


if __name__ == "__main__":
    app_root = tk.Tk()
    app_root.title("重量分析ツール")
    app_root.protocol("WM_DELETE_WINDOW", on_closing)

    btn = tk.Button(app_root, text="CSV選択して解析", command=run, height=2, width=30)
    btn.pack(pady=20)

    if len(sys.argv) > 1:
        csv_file_to_process = sys.argv[1]
        app_root.after(500, lambda: process_file(csv_file_to_process))

    app_root.mainloop()
